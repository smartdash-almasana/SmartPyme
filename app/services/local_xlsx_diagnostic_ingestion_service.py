from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from app.contracts.bem_payloads import BemSourceMetadata, CuratedEvidenceRecord, EvidenceKind
from app.repositories.curated_evidence_repository import CuratedEvidenceRepositoryBackend

_REQUIRED_COLUMNS = (
    "evidence_id",
    "precio_venta",
    "costo_unitario",
    "stock_actual",
    "ventas_periodo",
    "compras_periodo",
    "precio_lista",
    "costo_actual",
)


def _require_non_empty(value: str, field_name: str) -> None:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{field_name} es obligatorio y no puede estar vacio")


def _to_iso(value: datetime) -> str:
    if not isinstance(value, datetime):
        raise TypeError("now_provider debe retornar datetime")
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    else:
        value = value.astimezone(timezone.utc)
    return value.isoformat()


def _require_number(value: object, field_name: str, row_number: int) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{field_name} invalido en fila {row_number}")
    return float(value)


class LocalXlsxDiagnosticIngestionService:
    """Ingesta local minima XLSX -> CuratedEvidenceRecord para diagnostico."""

    def __init__(
        self,
        repository: CuratedEvidenceRepositoryBackend,
        now_provider: Callable[[], datetime] | None = None,
    ) -> None:
        self._repository = repository
        self._now_provider = now_provider or (lambda: datetime.now(timezone.utc))

    def ingest_file(self, tenant_id: str, file_path: str | Path) -> int:
        _require_non_empty(tenant_id, "tenant_id")
        path = Path(file_path)
        if not path.exists() or not path.is_file():
            raise ValueError("file_path no existe")
        if path.suffix.lower() != ".xlsx":
            raise ValueError("Solo se aceptan archivos .xlsx")

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if headers is None:
                raise ValueError("XLSX vacio")

            normalized_headers = [
                str(cell).strip().lower() if cell is not None else "" for cell in headers
            ]
            index_by_column = {name: idx for idx, name in enumerate(normalized_headers) if name}
            missing = [col for col in _REQUIRED_COLUMNS if col not in index_by_column]
            if missing:
                raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")

            imported = 0
            received_at = _to_iso(self._now_provider())
            source = BemSourceMetadata(source_name=path.name, source_type="xlsx_local")

            for row_number, row in enumerate(rows, start=2):
                if row is None or all(cell is None for cell in row):
                    continue

                raw_evidence_id = row[index_by_column["evidence_id"]]
                if not isinstance(raw_evidence_id, str) or not raw_evidence_id.strip():
                    raise ValueError(f"evidence_id invalido en fila {row_number}")

                payload = {
                    "precio_venta": _require_number(
                        row[index_by_column["precio_venta"]], "precio_venta", row_number
                    ),
                    "costo_unitario": _require_number(
                        row[index_by_column["costo_unitario"]], "costo_unitario", row_number
                    ),
                    "stock_actual": _require_number(
                        row[index_by_column["stock_actual"]], "stock_actual", row_number
                    ),
                    "ventas_periodo": _require_number(
                        row[index_by_column["ventas_periodo"]], "ventas_periodo", row_number
                    ),
                    "compras_periodo": _require_number(
                        row[index_by_column["compras_periodo"]], "compras_periodo", row_number
                    ),
                    "precio_lista": _require_number(
                        row[index_by_column["precio_lista"]], "precio_lista", row_number
                    ),
                    "costo_actual": _require_number(
                        row[index_by_column["costo_actual"]], "costo_actual", row_number
                    ),
                }

                record = CuratedEvidenceRecord(
                    tenant_id=tenant_id.strip(),
                    evidence_id=raw_evidence_id.strip(),
                    kind=EvidenceKind.EXCEL,
                    payload=payload,
                    source_metadata=source,
                    received_at=received_at,
                )
                self._repository.create(record)
                imported += 1

            if imported == 0:
                raise ValueError("XLSX sin filas de datos")
            return imported
        finally:
            workbook.close()
